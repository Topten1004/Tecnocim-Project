import datetime
import json
import pandas as pd
import numpy as np
import os
from django.shortcuts import render
from more_itertools import locate
from openpyxl.styles import PatternFill
from rest_framework.response import Response
from rest_framework import status
import sys
from openpyxl import load_workbook

from sources.functions import *
from core.models import *
from equivalencias.models import *

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
desired_width = 2500
pd.set_option('display.width', desired_width)

# columnas del documento Cirbe de BdE
# columns_names = ['Entidad', 'Operacion', 'Tipo', 'NatInterv', 'Participantes', 'Moneda', 'Plazo', 'GarantiaReal',
#                 'GarantiaPersonal', 'SituOper', 'Dispuesto', 'Importes', 'DemoraGastos', 'Disponible']

columns_names = ['entidad', 'operacion', 'tipo', 'natInterv', 'participantes', 'moneda', 'plazo', 'real',
                 'personal', 'situOper', 'dispuesto', 'importes', 'demora', 'disponible']

# CIF = 'ESB50560440'

"""
La función header se puede usar para eliminar todas las filas que no sean tabla y poseriormente se pueden
ordenar de acuerdo con su orden por columnas, lo cual permitiría construir el df por filas, pero resulta que
el cambio de página puede traer problemas porque cuando hay más de dos elementos en garantías personales
éstos no quedarían uno tras otro, sino ya tras el disponible. Es un problema con solución, de manera que si
la clasificación por posición se pusiera compleja, ésta sería una buena alternativa, ya que es invulnerable
a cambios de coordenadas:
    df['Row'] = [header(element) for element in df.Path]
    df = df[df.Row]
    df.sort_values(by=['Path'])
"""


def header(element):
    finalElements = []
    elements = element.split('/')
    for element in elements:
        finalElements += element.split('[')
    #print(finalElements)
    return not ('TH' in finalElements or 'TR' not in finalElements or 'TD' not in finalElements)


# no se usa
def joinGarantias(element):
    #print(element.GarantiaReal_descripcion)
    try:
        check = type(element.GarantiaReal_descripcion) == str
        if check:
            if len(element.GarantiaReal_descripcion) > 0:
                if type(element.GarantiaPersonal_descripcion) == str:
                    if len(element.GarantiaPersonal_descripcion) > 0:
                        return element.GarantiaReal_descripcion + ' / ' + element.GarantiaPersonal_descripcion
                    else:
                        return element.GarantiaReal_descripcion
            else:
                return element.GarantiaPersonal_descripcion
        else:
            try:
                check = type(element.GarantiaPersonal_descripcion) == str
                if check:
                    if len(element.GarantiaPersonal_descripcion) > 0:
                        return ' / '.join(
                            element.GarantiaReal_descripcion) + ' / ' + element.GarantiaPersonal_descripcion
                    else:
                        return element.GarantiaReal_descripcion
                else:
                    return ' / '.join(
                        element.GarantiaReal_descripcion) + ' / ' + ' / ' + element.GarantiaPersonal_descripcion
            except:
                return element.GarantiaReal_descripcion

    except:
        try:
            return element.GarantiaPersonal_descripcion
        except:
            return None


# a partir de la posición de un elemento del pdf, devolver el label de la columna que le corresponde en el dataframe
def position(entidades, element, extraccion):
    if element.B1 == 41:
        if element.Text.lower().split(',')[0] in entidades:
            # print('Nuevo Banco: ', element.Text)
            columna = columns_names[0]
        elif len(element.Text) < 4:
            if element.B3 == 509:
                columna = columns_names[8]
            elif element.B3 == 400 or element.B3 == 404 or element.B3 == 396:
                columna = columns_names[4]
            else:
                columna = columns_names[7]
        else:
            # print('Nueva Operación', element.Text)
            columna = columns_names[1]
    elif element.B1 == 307:
        columna = columns_names[2]
    elif element.B1 == 332:
        columna = columns_names[3]
    elif element.B1 == 364:
        columna = columns_names[4]
    elif element.B1 == 408:
        columna = columns_names[5]
    elif element.B1 == 437:
        columna = columns_names[6]
    elif element.B1 == 465:
        columna = columns_names[7]
    elif element.B1 == 493:
        columna = columns_names[8]
    elif 495 < element.B1 < 520:
        columna = columns_names[9]
    elif element.B3 == 614:
        columna = columns_names[10]
    elif element.B3 == 683:
        columna = columns_names[11]
    elif element.B3 == 752:
        columna = columns_names[12]
    elif element.B3 == 816:
        columna = columns_names[13]
    else:
        # este caso no se ha de dar nunca, si se da hay que revisar qué está fallando
        columna = columns_names[1]
        extraccion_error = Extracciones_Errores(**{'extraccion': extraccion,
                                                   'mensaje': f'Corrección susceptible de errores: operación {element.Text}',
                                                   'bloqueo': 3})
        extraccion_error.save()
        #print('No tendría que haber pasado: ', element.B1, ' ', element.Text)
    # print(columna)
    return columna


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.join(BASE_DIR, 'sources')

    return os.path.join(base_path, relative_path)


def entidadesFinancieras():
    # carga de los datos de las entidades financieras a partir del fichero xls del Banco de España
    # otra opción es recoger el contenido de la tabla ENTIDAD
    file_path = resource_path('resources') + '/REGBANESP_CONESTAB_A.xls'
    df = pd.read_excel(file_path, usecols=['COD_BE', 'NOMBRE50'],
                       dtype={'COD_BE': str, 'NOMBRE50': str})  # , encoding='utf-8')

    # Lista con los strings finales de los nombres compuestos de cada entidad financiera
    endings = []
    for index, x in df.iterrows():
        final = x.NOMBRE50.split(',')[-1]
        if final.strip() not in endings and len(x.NOMBRE50.split(',')) > 1:
            endings.append(final.strip())

    # Se añade a la lista de strings finales el caracter vacío, para propósitos de filtrado
    endings.append('')
    endings.append('SUCURSAL')
    endings.append('EN')
    endings.append('ESPAÑA')
    # print(endings)

    # Se cambian los nombres de los bancos a minúscula y se conserva solo el primer tramo del nombre, para simplificar comparaciones
    df['nombre50'] = [x.lower().split(',')[0] for x in df.NOMBRE50]
    df.NOMBRE50 = [x.strip() for x in df.NOMBRE50]

    # Se eliminan espacios en blanco de los códigos, para evitar errores en la comparación con los del Cirbe
    df.COD_BE = df.COD_BE.str.strip()

    # Se compone código+nombre a partir de los datos del BdE, para su comparación equivalente con la denominación de cada
    # banco en el Cirbe
    entidades = []
    for index, x in df.iterrows():  # print(x.CodigoSupervisor)
        entidades.append(x.COD_BE + ' ' + x.nombre50)

    entidades_completas = []
    for index, x in df.iterrows():  # print(x.CodigoSupervisor)
        entidades_completas.append(x.COD_BE + ' ' + x.NOMBRE50)

    return entidades, endings, entidades_completas

# quitar la cabecera de los datos, para tratar solo los datos relevantes
def filtrar_cabecera(df, entidades, extraccion):

    first = None
    # ordenar de manera que quede en el mismo orden que el Cirbe, de derecha a izquierda y de arriba a abajo
    df2 = df.loc[:, 'Page':'B4'].sort_values(by=['Page', 'B2', 'B1'], ascending=[True, False, True])
    # eliminar espacios en blanco para no tener sorpresas con las comparaciones
    df2['Text'] = df['Text'].str.strip()
    for index, x in df2.iterrows():
        if x.Text.lower().split(',')[0].strip() in entidades:
            first = index
            break
    if first == None:
        extraccion_error = Extracciones_Errores(**{'extraccion': extraccion,
                                                   'mensaje': 'No se ha detectado la primera línea de información bancaria. '
                                                              'Revisar nombres entidades', 'bloqueo': 1})
        extraccion_error.save()
        return pd.DataFrame()

    return df2.loc[first:, :]


def revisionEntidades(df2, entidades, endings, entidades_completas):
    data = []
    new_row = {}
    # analizamos fila a fila debido a errores de la API de Adobe, que puede juntar un banco con la primera operación
    for index, x in df2.iterrows():
        # si se trata de una entidad pero su última componente (ej. S.A.) no se corresponde con un valor válido
        # quiere decir que se ha fusionado el banco con la operación y es necesario separarlos y crear un registro adicional
        Text = x.Text.lower().split(',')
        TEXT = x.Text.split(',')
        # print(TEXT[-1].strip())
        if Text[0].strip() in entidades and (len(TEXT) > 1 and TEXT[-1].strip() not in endings):
            # seleccionar el texto tras la primera coma
            final = TEXT[-1].strip()
            # se puede tratar de una secuencia separada por espacios, y el código de operación puede estar también
            # separado por espacios y es necesario conservar todo el código. Así, se avanza trozo a trozo de la lista generada
            # cuando se identifica que un elemento de la lista se corresponde con el final de un nombre (ej. SCC), entonces
            # se puede asegurar que el resto es el código, y es lo que se conserva para el registro de la operación
            # print(final.split())
            loop = final.split()
            final = loop
            operation = ''
            for element in loop:
                # print(element)
                if element in endings:
                    final = final[1:]
                    # print('Final: ', final)
                else:
                    operation = ' '.join(final).strip()
                    # print('Operation: ', operation)
                    break
            # print(operation)
            # en el registro de la operación, se toma el nombre de la entidad bancaria que constaba en el principio del
            # texto compuesto por banco+operación
            entity = TEXT[0].strip().split()[0]
            entity = ''.join([element if entity in element else '' for element in entidades_completas])
            new_row['Text'] = entity
            # para el resto de los valores del registro (Page, B1-B4), se conservan los del registro original
            for i in df2.columns[2:]:
                new_row[i] = x[i]

            # print(new_row)
            # se añade el nuevo registro a la lista de registros a partir de la cual se construirá el df
            data.append(new_row)
            # se vacía el registro
            new_row = {}
            # se actualiza el texto al código de la operación
            x.Text = operation
            # se añade el registro orginal, con el texto actualizado al código de la operación, a la lista de registros
            data.append(x.to_dict())
        # si no había ningún problema con el nombre de la entidad, se añade el registro a la lista de registros
        else:
            data.append(df2.loc[index].to_dict())

    # se reconstruye el df con la lista de diccionarios (registros) que se ha construido con la revisión
    return pd.DataFrame(data)

    # se redondean las coordenadas a int para no tener problemas de precisión float cuando se comparen posiciones

# extraer el CIF y fecha del pdf; no se usa
def identidad(df):
    period = df[df.B1 == 37.82000732421875].iloc[0].Text
    period = period.split()[-1].split('-')

    CIF = df[df.B1 == 38.31050109863281].iloc[0].Text

    if len(CIF.split('/')) > 1:
        CIF = CIF.split('/')[0].split()[-3]
    else:
        CIF = CIF.split()[-1]

    return CIF, period[0], period[1]


def construirCirbe(df2, entidades, extraccion):
    # se genera un df vacío con las columnas del Cirbe

    # se van a construir registros haciendo un barrido de los registros del df df2 de principio a fin;
    # la primera posición siempre es una entidad bancaria y la segunda la primera operación, de manera que
    # se inicia la construcción del df rellenando esas dos primera posiciones del primer registro
    data = []
    new_row = {}
    entidad = df2.Text[0]
    operacion = df2.Text[1]
    new_row['entidad'] = entidad
    new_row['operacion'] = operacion

    # cuando se detecte una nueva entidad bancaria, habrá que guardar el registro anterior; también cuando se detecte una
    # nueva operación, excepto en el caso de la primera operación de una nueva entidad bancaria, para lo que usamos
    # el siguiente indicador:
    first_operation = False

    # barrer los registros para ir construyendo un nuevo df estilo Cirbe
    for index, x in df2.iloc[2:, :].iterrows():
        # print(x)
        # ver el label que le corresponde al registro (ej. Entidad, Disponible, Garantía, etc)
        columna = position(entidades, x, extraccion)
        # print(columna)
        # si se trata de una entidad
        if columna == columns_names[0]:
            # salvar el registro construido previamente
            data.append(new_row)
            # print(new_row)
            # actualizar el nombre de la entidad bancaria a la nueva entidad
            entidad = x.Text
            # vaciar el registro para iniciar la construcción de un nuevo registro
            new_row = {}
            # añadir, como primer elemento del registro, el nombre de la nueva entidad bancaria
            new_row[columns_names[0]] = entidad
            # cambiar el flag para que no salve estos primeros datos incipientes cuando, seguidamente, detecte la
            # primera operación de esta nueva entidad bancaria
            first_operation = True
        # si se trata de una operación
        elif columna == columns_names[1]:
            # si se trata de la primera operación de la entidad, asignar el código al campo de operación
            if first_operation:
                new_row[columns_names[1]] = x.Text
                # cambiar el flag para que, en caso de una nueva operación, se salve el registro
                first_operation = False
            # si no se trata de una primera operación, salvar el registro construido para la operación anterior
            else:
                data.append(new_row)
                # print(new_row)
                # vaciar el registro para iniciar la construcción del nuevo registro de la nueva operación
                new_row = {}
                # asignar valor al primer campo - entidad bancaria
                new_row[columns_names[0]] = entidad
                # asignar valor al segundo campo - operación
                new_row[columns_names[1]] = x.Text
        # si se trata de otro campo
        else:
            # si ya existía un registro para ese campo (ej. Garantías), generar una lista con los registros anteriores y el nuevo
            if columna in new_row.keys():
                # new_row[columna] = add_to_columna(new_row[columna], x.Text)
                new_row[columna] += ("," + x.Text)
            # si no existía ese campo todavía, generarlo y asignarle el valor que le pertoca
            else:
                new_row[columna] = x.Text

    # print(new_row)
    # añadir a la lista de diccionarios el último registro generado, que todavía no se ha salvado
    data.append(new_row)

    # crear el df cirbe con la lista de diccionarios construida
    return pd.DataFrame(data, columns=columns_names)


def salvarXLSX(cirbe, file, sheetName):
    if os.path.isfile(file):
        os.remove(file)

    cirbe.to_excel(file, sheet_name=sheetName, index=False)
    return True


def producto_del_pool(cirbe):

    # re-clasificar instrumentos financieros de acuerdo con los criterios de Alia
    cirbe['producto'] = cirbe.apply(lambda x: poolBancario(int(x.entidad.split()[0]), x.tipo, x.plazo, x.disponible, x.dispuesto), axis=1)

    # eliminar elementos del pool que hayan sido clasificados como None
    new_cirbe = cirbe[cirbe.producto != None]

    columns_names = ['entidad', 'operacion', 'producto', 'tipo', 'natInterv', 'solCol', 'participantes', 'moneda', 'plazo',
                     'real', 'personal', 'situOper', 'dispuesto', 'importes', 'demora', 'disponible']

    # reordenar las columnas de acuerdo con la presentación estándar del Cirbe
    new_cirbe = new_cirbe[columns_names]

    return new_cirbe

# comprobar que los elementos del cirbe no cumplen un conjunto de condiciones básicas; su cumplimiento indica un error
def check(cirbe):
    #print(cirbe.columns)
    #for index, x in cirbe[['operacion', 'tipo', 'moneda', 'plazo', 'dispuesto', 'importes', 'demora', 'disponible']].iterrows():
        #print(x.operacion, x.tipo, x.moneda, x.plazo)
    error = [
        True if len(x.operacion.split(',')) > 1 or len(x.tipo.split(',')) > 1 or len(x.moneda.split(',')) > 1 or
                len(x.plazo.split(',')) > 1 or x.dispuesto == np.nan or x.disponible == np.nan
        else False for index, x in cirbe[['operacion', 'tipo', 'moneda', 'plazo', 'dispuesto', 'importes', 'demora', 'disponible']].iterrows()]

    if True in error:
        filas = list(locate(error, lambda a: a == True))
        filas = [a + 2 for a in filas]
        return True, filas

    else:
        return False, None

# por ahora no se usa mientras se determina cómo se quiere relacionar el campo producto del cirbe con el campo producto del contrato
def arrangeRow(register, row):
    if not register.contrato.producto:
        register.contrato.producto = row.producto
        register.save()
    row.drop(['producto'], inplace=True)
    return row

# comprobar que los datos de cada elemento de la row del df cumplen una condiciones básicas; el caso contrario es un error
def row_check(row):
    #print(row)
    type_check = {}
    type_check['entidad'] = row['entidad'].split()[0] in list(Entidad.objects.all().values_list('codigo', flat=True))
    type_check['operacion'] = row['operacion'] != None
    type_check['tipo'] = row['tipo'] in list(Tipo.objects.all().values_list('tipo', flat=True))
    type_check['natInterv'] = row['natInterv'] in list(NatInterv.objects.all().values_list('tipo', flat=True))
    if row['solCol']:
        type_check['solCol'] = row['solCol'] in list(SolCol.objects.all().values_list('tipo', flat=True))
        type_check['participantes'] = type(row['participantes']) == int
    type_check['moneda'] = row['moneda'] in list(Moneda.objects.all().values_list('tipo', flat=True))
    type_check['plazo'] = row['plazo'] in list(Plazo.objects.all().values_list('tipo', flat=True))

    if row['real']:
        garantias = list(Real.objects.all().values_list('tipo', flat=True))
        type_check['real'] = True
        for element in row['real'].split(','):
            if element not in garantias:
                type_check['real'] = False
                break

    if row['personal']:
        garantias = list(Personal.objects.all().values_list('tipo', flat=True))
        type_check['personal'] = True
        for element in row['personal'].split(','):
            if element not in garantias:
                type_check['real'] = False
                break

    if row['situOper']:
        #print('SituOper: ', list(SituOper.objects.all().values_list('tipo', flat=True)))
        type_check['situOper'] = row['situOper'] in list(SituOper.objects.all().values_list('tipo', flat=True))

    type_check['dispuesto'] = type(row['dispuesto']) == float
    type_check['importes'] = type(row['importes']) == float or None
    type_check['demora'] = type(row['demora']) == float or None
    type_check['disponible'] = type(row['disponible']) == float

    return type_check


# rutina para salvar un registro que gestiona los campos manytomany de las garantías
# todavía está por acabar de incorporar try-except para gestionar posibles errores
def saveRow(row):

    garantiaReal = None
    garantiaPersonal = None
    if row.real:
        garantiaReal = row.real.split(',')
        #print(f'Garantia Real: {garantiaReal}')
    if row.personal:
        garantiaPersonal = row.personal.split(',')
        #print(f'Garantia Personal: {garantiaPersonal}')

    row.drop(['real', 'personal'], inplace=True)
    #pool = Pool.objects.filter(contrato=row.contrato, documento__fecha=row.documento.fecha)
    #if pool:
    #    dispuesto = pool.aggregate(Sum('dispuesto'))
    #    if dispuesto != row.dispuesto:
    #        #habría de ser un error pero por ahora ilustramos, no paramos
    #        Extracciones_Errores(
    #            {'extraccion': row.extraccion, 'mensaje': f'Error: dispuesto del Cirbe {row.dispuesto} no se corresponde '
    #                                                      f'con el dispuesto del BSS {dispuesto}', 'bloqueo': 3,
    #             'tabla': 'Cirbe', 'campo': 'dispuesto'}).save()
    #print('Row: ', row)
    data = Cirbe(**row)
    data.save()

    if garantiaReal:
        for element in garantiaReal:
            #print(element)
            if element:
                data.real.add(Real.objects.get(tipo=element))

    if garantiaPersonal:
        for element in garantiaPersonal:
            #print(element)
            if element:
                data.personal.add(Personal.objects.get(tipo=element))

    return data

# pasar una fila del dataframe con primitivas a una fila con instancias y con la que poder generar un registro
# no se hace justo después de completar el dataframe porque están las rutinas de comprobación de errores
def transform_row_to_models(row, extraccion):

    try:
        row['entidad'] = Entidad.objects.get(codigo=row.entidad.split(' ')[0])
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Entidad {row.entidad} no existe, revisar y actualizar entidades (administrador'
                                                     f' base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False

    try:
        row['tipo'] = Tipo.objects.get(tipo=row['tipo'])
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Tipo {row.tipo} no existe, revisar y actualizar Tipo (administrador base de datos)',
                'traza': f'{e}', 'bloqueo': 1}, False
    try:
        row['producto'] = Producto.objects.get(tipo=row['producto'])
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Producto {row.producto} no existe, revisar y actualizar productos '
                                                     f'(administrador base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False
    if row['solCol']:
        try:
            row['solCol'] = SolCol.objects.get(tipo=row['solCol'])
        except Exception as e:
            return {'extraccion': extraccion, 'mensaje': f'SolCol {row.solCol} no existe, revisar y actualizar solCol '
                                                         f'(administrador base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False
    else:
        row.drop(['solCol'], inplace=True)

    try:
        row['natInterv'] = NatInterv.objects.get(tipo=row['natInterv'])
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'NatInterv {row.natInterv} no existe, revisar y actualizar natInterv '
                                                     f'(administrador base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False

    try:
        row['moneda'] = Moneda.objects.get(tipo=row['moneda'])
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Moneda {row.moneda} no existe, revisar y actualizar monedas '
                                                     f'(administrador base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False

    if row['situOper']:
        try:
            row['situOper'] = SituOper.objects.get(tipo=row['situOper'])
        except Exception as e:
            return {'extraccion': extraccion, 'mensaje': f'SituOper {row.situOper} no existe, revisar y actualizar situOper '
                                                         f'(administrador base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False
    else:
        row.drop(['situOper'], inplace=True)

    try:
        row['plazo'] = Plazo.objects.get(tipo=row['plazo'])
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Plazo {row.plazo} no existe, revisar y actualizar plazos (administrador '
                                                     f'base de datos)', 'traza': f'{e}', 'bloqueo': 1}, False
    return row, True

def searchContract(entidad, operacion, fecha):
    contrato = None
    try:
        # coger el cirbe más nuevo de entre los similares al que estamos tratando
        register = Cirbe.objects.filter(entidad=entidad, operacion=operacion).exclude(contrato__isnull=True)
        #print('En searchContract: ', register)
        # si existe y el contrato está vigente, asignárselo al cirbe que estamos tratando
        for element in register:
            if element.contrato.inicio <= fecha <= element.contrato.vencimiento:
                contrato = element.contrato
                break
    except:
        pass

    return contrato

def saveCirbe(row, doc, extraccion):
    # el dataframe tiene primitivas que se usan para generar instancias de modelos
    row, success = transform_row_to_models(row, extraccion)

    # si se produce un error, el elemento devuelto no será una fila sinó una dupla
    if not success:
        #print(row)
        return row, success
    # añadir el campo documento a la fila, para que se corresponda con los parámetros de un modelo Cirbe
    row['documento'] = doc

    # mirar si ya existe ese registro y hacer un update si es así, conservando el link a contrato
    try:
        #print(f'Pruebo a ver si existe cirbe con {row.entidad} {row.operacion} {doc.fecha}')
        #print(Cirbe.objects.filter(entidad=row.entidad))
        register = Cirbe.objects.get(entidad=row['entidad'], operacion=row['operacion'], documento__fecha=doc.fecha)
        row['id'] = register.id
        if register.contrato: row['contrato'] = register.contrato
        else: row['contrato'] = searchContract(row.entidad, row.operacion, doc.fecha)

    # si no existe el registro, primero se mira si hay algún cirbe igual con contrato todavía vigente
    except:
        row['contrato'] = searchContract(row.entidad, row.operacion, doc.fecha)
    row['extraccion'] = extraccion
    saveRow(row)

    return row, success

# load json file and save table
def cirbe(file, doc, extraccion, i=1):
    # recuperar entidades financieras de bbdd o construirlas
    try:
        entidades, endings, entidades_completas = entidadesFinancieras()
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Fallo en entidadesFinancieras(); comprobar la existencia de '
                                                     f'{resource_path("resources") + " / REGBANESP_CONESTAB_A.xls"}',
                'traza': f'{e}', 'bloqueo': 1}, False
    try:
        # leer el json generado por acrobat sdk
        with open(file, encoding='utf-8') as fh:
            data = json.load(fh)
        fh.close()

        # seleccionar solo aquellos campos que nos interesan
        elements = data['elements']
        keep = ['Path', 'Bounds', 'Page', 'Text']

        # deshacerse de los valores nulos
        df = pd.DataFrame.from_records(elements)[keep].dropna()
        # print(df.Text)
        # generar 4 columnas a partir de la lista de coordenadas
        # pd.DataFrame(df['Bounds'].to_list(), columns=['B1','B2','B3','B4'])
        df[['B1', 'B2', 'B3', 'B4']] = pd.DataFrame(df.Bounds.tolist(), index=df.index)
        # print(df[df.B1 == 37.82000732421875].Text)

        # identificar esta información del doc; ahora no se usa porque la proporciona el usuario, pero a futuro puede resultar útil
        CIF, mes, year = identidad(df)

        # conservar solo los campos de interés, ordenar de manera que los cabeceros queden al principio;
        # rehacer el índice y borrar el previo
        df.drop(labels=['Path'], axis=1, inplace=True)
        df = df.loc[:, 'Page':'B4'].sort_values(by=['B2', 'Page'], ascending=[False, True]).reset_index()
        # df = df.loc[:, 'Page':'B4'].sort_values(by=['Page', 'B2', 'B1'], ascending=[True, False, True])
        df['Text'] = df['Text'].str.strip()
        for column in ['B1', 'B2', 'B3', 'B4']:
            df[column] = [round(x) for x in df[column]]
        df = df[df.B2 < 439]
        df = df[df.B1 != 38]
        # df.drop(columns = ['index'], inplace = True)
        # separar banco de operación cuando el procesamiento del pdf los ha juntado
        #print(df)
        df = revisionEntidades(df, entidades, endings, entidades_completas)
        # identificar hasta dónde llegan los registros con cabezales y borrarlos
        df = filtrar_cabecera(df, entidades, extraccion)
        if df.empty:
            return 'Fallo filtrando cabecera', False
        df.sort_values(by=['Page', 'B2', 'B3'], ascending=[True, False, True], inplace=True)
        df.reset_index(inplace=True)
        last = df[df.Text == 'TOTAL EN EL SISTEMA'].index.tolist()[0]
        df2 = df.loc[:last - 1, :]
        # df2.sort_values(by=['index'], ascending=[True], inplace=True)
        # se eliminan los registros de comentarios y se rehace el índice y se elimina el previo
        # df = df.reset_index()
        df2 = df2.drop(columns=['index'])
        # se identifica dónde empiezan los registros con los comentarios finales y se eliminan

        # df2.sort_values(by=['Page', 'B4'], ascending=[True, False], inplace=True)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Fallo en procesamiento del json antes de empezar la interpretación. '
                                                     f'Comprobar que el json {file} proviene de un Cirbe.',
                'traza': f'{e}', 'bloqueo': 1}, False
    #print(df2)
    try:
        cirbe = construirCirbe(df2, entidades, extraccion)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Fallo en construirCirbe(), función de interpretación del json; '
                                                     f'comprobar que el json {file} proviene de un Cirbe.',
                'traza': f'{e}', 'bloqueo': 1}, False

    # consultar el nombre de las columnas del df cirbe
    columns = cirbe.columns

    # si hay columnas que un Cirbe tiene pero que no hallamos en el df, creamos los campos correspondientes
    cirbe[Diff(columns_names, columns)] = np.nan

    # los códigos de operaciones son muy complejos y se pueden haber dividido en varias partes, dando lugar a filas cuyo
    # único elemento es el código de operación, estando el resto de campos vacíos. La siguiente rutina agrega todas esas
    # filas de una operación común en una única fila, conservando los campos válidos y construyendo un código de operación
    # con la concatenación de todos los códigos disgregados
    try:
        cirbe = unificar_operaciones(cirbe)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': 'Error en la revisión del cirbe para reunificar códigos de operaciones que acrobat '
                                                     'haya separado (unificar_operaciones()).', 'traza': f'{e}', 'bloqueo': 1}, False
    # separar la información de productos colaborativos en código y número de participantes
    try:
        cirbe['solCol'] = cirbe.participantes.apply(lambda x: sol_col(x))
        cirbe['participantes'] = cirbe.participantes.apply(lambda x: num_participantes(x))
        # asegurar que no hay nans ni strings vacíos en personal y real
        cirbe['personal'] = cirbe.personal.apply(lambda x: personal(x))
        cirbe['real'] = cirbe.real.apply(lambda x: real(x))
        # asegurar que no hay nans en situOper
        cirbe['situOper'] = cirbe.situOper.replace({np.nan:None})
        # transformar de string a float los campos cuantitativos
        cirbe['disponible'] = [changeToFloat(i, x.disponible, extraccion, 'disponible') for i, x in cirbe.iterrows()]
        cirbe['dispuesto'] = [changeToFloat(i, x.dispuesto, extraccion, 'dispuesto') for i, x in cirbe.iterrows()]
        cirbe['importes'] = [changeToFloat(i, x.importes, extraccion, 'importes') for i, x in cirbe.iterrows()]
        cirbe['demora'] = [changeToFloat(i, x.demora, extraccion, 'demora') for i, x in cirbe.iterrows()]
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': 'Error en el proceso de normalización de los datos del Cirbe - operación '
                                                     'tras unificar_operaciones()', 'traza': f'{e}', 'bloqueo': 1}, False
    # comprobar posibles errores por fusión de celdas del intérprete de pdfs de acrobat
    try:
        status, fila = check(cirbe)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': 'Error en el chequeo de errores tras la normalización - función check()',
                'traza': f'{e}', 'bloqueo': 1}, False
    #print(status, fila)

    # clasificar productos de acuerdo con los criterios de Alia;
    # posteriormente se podrá comparar con la información del contrato para comprobar la consistencia
    try:
        cirbe = producto_del_pool(cirbe)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': 'Error en el proceso de clasificación de operaciones bancarias: función '
                                                     'producto_del_pool()', 'traza': f'{e}', 'bloqueo': 1}, False
    # el disponible no tiene sentido, ya que se calcula con el límite;
    # se podría conservar para comprobar la consistencia también
    # cirbe.drop(['disponible'], axis=1, inplace=True)

    # comprobar que la naturaleza de la información contenida en cada celda es la que le corresponde
    # los False en type_check serán errores
    type_check = []
    # generar un registro en la BBDD a partir de cada fila del dataframe
    try:
        for i, row in cirbe.iterrows():
            row.dispuesto = abs(round(row.dispuesto, 2))
            type_check.append(row_check(row))
            row, success = saveCirbe(row, doc, extraccion)
            if not success:
                return row, success
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': f'Fallo en la rutina de salvar cirbe a la bbdd', 'traza': f'{e}', 'bloqueo': 1}, False

    doc.status = True
    doc.save()

    # informar sobre posibles errores a partir de type_check;
    # errores en type_check también se devuelven como resultado de la api
    errores = []
    for element in type_check:
        if False in element.values():
            #print(f'Posible error en elemento: {element.items()}')
            extraccion_error = Extracciones_Errores(**{'extraccion': extraccion,
                                                       'mensaje': f'Posible error en {element.items()}', 'bloqueo': 3})
            extraccion_error.save()
            errores.append(element)

    # salvar resultados en un excel automáticamente; se puede eliminar o separar como api independiente
    """
    file = os.path.join(MEDIA_ROOT, os.path.splitext(doc.documento.name)[0]+'.xlsx')
    if salvarXLSX(cirbe, file, 'raw'): print('Salvado fichero ', file, ' Número ', i, ' de la secuencia')

    if status:
        print(f'\nPOSIBLES ERRORES EN FILA/S {fila}, SE RECOMIENDA EDICIÓN MANUAL\n')
        wb = load_workbook(file)
        ws = wb["raw"]
        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')
        for row in fila:
            cell = 'A' + str(row)
            ws[cell].fill = redFill
        wb.save(file)
        wb.close()
    """
    #print(type_check)
    #print(f'\nNúmero de instrumentos financieros: {len(cirbe)}\n')

    return errores, True


if __name__ == '__main__':
    # status, cirbeDF = cirbe('C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\Cirbe\\hecho\\60- CIRBE 311221.json')
    # status, cirbeDF = cirbe('C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\Cirbe\\hecho\\Atipico Cirbe Informe detallada a septiembre 2020.json')
    status, cirbeDF = cirbe('C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\Cirbe\\hecho\\CIRBE FCO SIMO SL.json')

    print(cirbeDF)
    # saveToMSSQL(cirbeDF)
    # cirbe('C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\Cirbe\\hecho\\Atipico Cirbe Informe detallada a septiembre 2020.json', 1)
    # cirbe('C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\Cirbe\\hecho\\60- CIRBE 311221.json', 1)
