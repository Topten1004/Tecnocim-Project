import json
# import logging

import xlsxwriter
import openpyxl
import pandas as pd
import os

import xlrd

from app import settings
from app.settings import MEDIA_ROOT
from core import dictionaries
from core.models import Empresa, Documento, Extracciones_Errores, Crudos


# salvar el excel original en un dataframe y nuevamente en un excel;
# es un método de deshacer celdas fusionadas y asegurar una lectura posterior del excel sin errores

def preProcess(excel, extraccion):

    if settings.EXTRACCION:
        excel = extraccion.ruta
        new_name = os.path.join(os.path.dirname(excel), os.path.basename(excel).rsplit('.', 1)[0] + '_unmerged.xlsx')
    else:
        new_name = os.path.join(MEDIA_ROOT, os.path.dirname(excel), os.path.basename(excel).rsplit('.', 1)[0] + '_unmerged.xlsx')
        #print('New name: ', new_name)
        excel = os.path.join(MEDIA_ROOT, excel)
        #print('Excel: ', excel)

    filename, file_extension = os.path.splitext(excel)
    # logging.info(filename)
    # logging.info(file_extension)

    if file_extension == '.xls':
        ### archivos con software de terceros puede llegar con codificaciones atípicas, habría que hacer try except para
        ### explorar distintas codificaciones con encoding_override a ver si se da con la correcta
        book = xlrd.open_workbook(excel)
        writer = pd.ExcelWriter(new_name, engine='xlsxwriter')
        for sh in book.sheet_names():
            df = pd.read_excel(excel, sheet_name=sh, header=None)
            df.to_excel(writer, sheet_name=sh, index=False, header=False)
        writer.save()
        book.release_resources()
    else:
        book = openpyxl.load_workbook(excel)
        writer = pd.ExcelWriter(new_name, engine='xlsxwriter')
        for sh in book:
            df = pd.read_excel(excel, sheet_name=sh.title, header=None)
            df.to_excel(writer, sheet_name=sh.title, index=False, header=False)
        writer.save()
        book.close()

    extraccion.ruta_unmerged = new_name
    extraccion.save()

    return new_name

# rutina para detectar en que coordenadas se encuentra la primera cuenta de la cuenta 100
def detect_first(book):

    coordenadas = {}
    hay_coordenadas = False
    for sh in book.worksheets:
        page = sh
        coordenadas['pagina'] = page.title
        for rx in range(1, sh.max_row+1):
            #print(sh.row(rx))
            for cx in range(1, sh.max_column+1):
                #print(f'row {rx} col {cx} value {sh.cell(rx, cx).value}')
                cell = sh.cell(rx, cx).value
                cell = str(cell)
                cell.replace('.', '')
                length = len(cell)
                #print(cell)
                if length>2:
                    tres_digitos = cell[:3]
                    if tres_digitos == '100':
                        coordenadas['100x'] = rx
                        coordenadas['cuentas'] = cx
                        hay_coordenadas = True
                        break
            if hay_coordenadas: break
        if hay_coordenadas: break

        #print(coordenadas)
    return coordenadas

# detectar en qué fila se encuentran los headers de la tabla, que son los que permiten idenficar la columna de saldos netos
# se hace una lectura inversa de las celdas a partir de las coordenadas de la cuenta 100 y se toma como fila de headers
# aquella que tenga un contenido no nulo y de tipo string
def detect_header(coordenadas, sh):
    fila_header = None
    for rx in range(coordenadas['100x'] - 1, 0, -1):
        cell = sh.cell(rx, coordenadas['cuentas']).value
        if cell!='':
            if type(cell) == str:
                try:
                    cell = cell.replace('.', '')
                    cell = cell.replace(',', '')
                    cell = float(cell)
                except:
                    fila_header = rx
                    break
    return fila_header

# lectura del excel a partir de las coordenadas presentes en el fichero del campo configFile de la empresa
def select_content(empresa, book):

    f = open(empresa.configFile)
    coordenadas = json.load(f)
    #print('Coordenadas en select: ', coordenadas)
    # df vacío a rellenar con los contenidos del excel
    df = pd.DataFrame(columns=['cuenta', 'concepto', 'magnitud'])
    df.flags.allows_duplicate_labels = False
    try:
        sh = book.get_sheet_by_name(coordenadas['pagina'])
    except:
        #print('Hay que configurar los parámetros de lectura del fichero Excel')
        return df, 0
    #columna = sh.col_values(coordenadas['columna'])[coordenadas['fila']:]
    columna = [sh.cell(row, coordenadas['cuentas']).value for row in range(1, sh.max_row+1)]
    columna = columna[coordenadas['100x']-1:]

    if str(columna[0])[0:3] != '100':
        #print('Hay que configurar los parámetros de lectura del fichero Excel')
        return df

    digitos = ['.','0','1','2','3','4','5','6','7','8','9']
    # determinar la longitud máxima de las cuentas
    # las que tengan una longitud menor son agregaciones que no se usarán
    max = 0
    length = 0
    for element in columna:
        cuenta = str(element)
        # si el número de cuenta se había fusionado con el concepto, hay que separarlos
        if 'separador' in coordenadas.keys():
            cuenta = cuenta.split(coordenadas['separador'])[0]
        # se normalizan las cuentas por si tienen puntos o separaciones entre grupos de cifras
        # se conserva el string original porque es el que más adelante marcará la longitud de las cuentas de interés
        if len(cuenta) > 3:
            cuenta_original = cuenta
            cuenta = cuenta.replace('.', '')
            cuenta = cuenta.replace(' ', '')
            try:
                int(cuenta)
                length = len(cuenta_original)
            # si no es convertible a int, quiere decir que es algún texto intermedio que mejor eliminar
            except:
                pass
                #columna.remove(element)
                #length = 0
            # print("Length: ", length)
            if length > max:
                max = length

    if max < 3:
        #print('Hay que configurar los parámetros de lectura del fichero Excel')
        return df

    # barrido de las columnas de interés
    for rx in range(coordenadas['100x'], sh.max_row+1):
        #print(len(str(cuenta)))
        saldo = sh.cell(rx, coordenadas['saldo']).value
        """
        if sh.cell_type(rx, columna_saldo) == xlrd.XL_CELL_EMPTY:
            if sh.cell_value(rx, columna_saldo-1) != saldo and sh.cell_type(rx, columna_saldo-2) != xlrd.XL_CELL_EMPTY:
                saldo = -1.0*sh.cell_value(rx, columna_saldo - 2)
        """
        if 'separador' in coordenadas.keys():
            cuenta = str(sh.cell(rx, coordenadas['cuentas']).value).split(coordenadas['separador'])[0].strip()
            codigo = str(sh.cell(rx, coordenadas['conceptos']).value).split(coordenadas['separador'])[1].strip()
        else:
            cuenta = str(sh.cell(rx, coordenadas['cuentas']).value).strip()
            codigo = str(sh.cell(rx, coordenadas['conceptos']).value).strip()
        if len(str(cuenta))==5 and str(cuenta)[:3] != str(sh.cell(rx+1, coordenadas['cuentas']).value)[:3] and isinstance(codigo, float):
            data = [cuenta, codigo, saldo]
            df.loc[len(df)] = data
        elif max-1 <= len(str(cuenta)) <= max and str(cuenta)[-1] in digitos:
            data = [cuenta, codigo, saldo]
            df.loc[len(df)] = data

    return df, max


def get_model_fields(model):
    return model._meta.fields

def detect_fields(doc, extraccion):

    #book = openpyxl.load_workbook(excel)
    excel = preProcess(doc.documento.name, extraccion)
    book = openpyxl.load_workbook(excel)
    configFile = os.path.join(MEDIA_ROOT, os.path.dirname(doc.documento.name), 'CoordenadasExcel.json')
    try:
        coordenadas = detect_first(book)
        if coordenadas:
            #print('Coordenadas: ', coordenadas)
            sh = book.get_sheet_by_name(coordenadas['pagina'])
            #print('Antes de detect_header')
            fila_headers = detect_header(coordenadas, sh=sh)
            if not fila_headers: return book
            #print('Fila Headers: ', fila_headers)
            # contenido de fila headers
            data = [sh.cell(fila_headers, col).value for col in range(1, sh.max_column + 1)]
            # eliminar celdas vacías desde el final de la fila
            while data[-1] == None: data.pop()
            # eliminar celdas que no contengan la palabra 'saldo', desde el final y hasta que se detecte la palabra 'saldo'
            while 'saldo' not in data[-1].lower(): data.pop()
            if data: coordenadas['saldo'] = len(data)
            # por defecto se asume que la columna con los conceptos va inmediatamente después de la de cuentas
            coordenadas['conceptos'] = coordenadas['cuentas'] + 1
            # print('Selección de filas y columnas: ', coordenadas)
            # respuesta = input('¿La selección es correcta (s/n)?: ').lower()
            # if 'n' in respuesta:
            #    coordenadas = selectFields(configFile)
            # else:
            #    coordenadas['conceptos'] = int(input('Número de columna con los conceptos: '))
            #    if coordenadas['conceptos'] == coordenadas['cuentas']:
            #        coordenadas['separador'] = input('Indique el separador entre cuenta y concepto: ')

            # se salvan las coordenadas en el fichero configFile que a su vez se salva en el directorio de
            # la empresa y se vincula a la empresa
            with open(configFile, 'w') as json_file:
                json.dump(coordenadas, json_file)
            doc.empresa.configFile = configFile
            doc.empresa.save()

        else:
            doc = False

    except:
        #print('La aplicación no ha podido detectar automáticamente las columnas del excel con la información.\n'
        #      'Por favor proceder a edición manual')
        doc = False

    return configFile, doc

# lectura del excel y generación de dataframe con la información estructurada

def read(doc, conceptos, extraccion):

    try:
        cuenta = Crudos.objects.get(documento=doc, concepto='correccion existencias activo')
        cuenta.delete()
    except:
        pass

    try:
        cuenta = Crudos.objects.get(documento=doc, concepto='correccion existencias resultados')
        cuenta.delete()
    except:
        pass

    #unMergeCell(excel_file, excel)
    try:
        excel = preProcess(doc.documento.name, extraccion)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': 'Fallo en el pre-procesamiento del excel (función preProcess). Posible archivo excel corrupto',
                'traza': f'{e}', 'bloqueo': 1}, False
    try:
        book = openpyxl.load_workbook(excel)
    except Exception as e:
        return {'extraccion': extraccion, 'mensaje': 'Fallo en la carga del excel pre-procesado', 'traza': f'{e}', 'bloqueo': 1}, False

    # conformar el futuro archivo configFile por si resulta necesario
    configFile = os.path.join(MEDIA_ROOT, os.path.dirname(doc.documento.name), 'CoordenadasExcel.json')

    # intentar recuperar fichero de configuración de lectura del excel
    empresa = Empresa.objects.get(CIF=doc.empresa.CIF)
    # ver si empresa tiene configFile con coordenadas de lectura del excel
    try:
        open(os.path.join(empresa.configFile))
    # si no la tiene, revisar si está en el directorio de la empresa para BSS
    except:
        #configFile = os.path.join(MEDIA_ROOT, os.path.dirname(doc.documento.name),
        #                          os.path.basename(doc.documento.name).rsplit('.', 1)[0] + '.json')
        # si el fichero existe, recuperar las coordenadas
        """
        if os.path.exists(configFile):
            try:
                f = open(configFile)
                json.load(f)
                empresa.configFile = configFile
                empresa.save()
            except Exception as e:
                return {'extraccion': extraccion, 'mensaje': f'Existe un fichero configFile pero es inválido. Fallo en función read() '
                                                             f'de XLSX.py con {empresa.configFile}', 'traza': f'{e}', 'bloqueo': 1}, False
        """
        # si no existe, aplicar las rutinas por defecto de detección de coordenadas
        #else:
        configFile, doc = detect_fields(doc, extraccion)
        if not doc:
            return configFile, 'Rutina detect_fields no ha podido detectar unas coordenadas excel válidas'

    while True:
        try:
            df, max = select_content(empresa, book)
        except:
            try:
                configFile, doc = detect_fields(doc, extraccion)
                if not doc:
                    return configFile, 'Rutina detect_fields no ha podido detectar unas coordenadas excel válidas'
                try:
                    df, max = select_content(empresa, book)
                except Exception as e:
                    return {'extraccion': extraccion, 'mensaje': 'Error rutina select_contents, coordenadas excel posiblemente inválidas',
                            'traza': f'{e}', 'bloqueo': 2}, False
            except Exception as e:
                return {'extraccion': extraccion, 'mensaje': 'Error rutina select_contents, coordenadas excel posiblemente inválidas',
                        'traza': f'{e}', 'bloqueo': 2}, False
        if df.empty:
            try:
                configFile, doc = detect_fields(doc, extraccion)
                df, max = select_content(empresa, book)
            except Exception as e:
                return {'extraccion': extraccion, 'mensaje': 'Nuevo error rutina select_contents, coordenadas excel posiblemente inválidas',
                        'traza': f'{e}', 'bloqueo': 2}, False
            if df.empty:
                return {'extraccion': extraccion, 'mensaje': 'Sigue devolviendo df.empty', 'bloqueo': 2}, False
            empresa.configFile = configFile
            empresa.save()

        try:
            df = df.astype({"magnitud": float})
            break
        except:
            newvalues = [a.replace(',', '') if type(a) == str else a for a in df.magnitud]
            try:
                newvalues = [float(a) for a in newvalues]
                df.magnitud = newvalues
                break
            except Exception as e:
                return {'extraccion': extraccion, 'mensaje': f'Excel aparentemente incorrecto. Revisar contenidos. '
                                                             f'Coordenadas utilizadas {empresa.configFile}',
                        'traza': f'{e}', 'bloqueo': 1}, False

    df['magnitud'] = df.magnitud.fillna(0)

    df['CuentaTemp'] = [str(cuenta)[0:3] for cuenta in df.cuenta]
    # se suman todas las cuentas que empiecen por la misma cifra que 'element'
    cantidad = df[df['CuentaTemp'] == '100']['magnitud'].sum()
    if cantidad < 0:
        sign = -1
    else:
        sign = 1

    #resultados = 0
    #for cuenta in dictionaries.conceptosAbreviado['PASIVO']['resultado del ejercicio']:
    #    df.CuentaTemp = df.cuenta[:len(cuenta)]
    #    resultados += df[df['CuentaTemp'] == str(cuenta)]['magnitud'].sum()

    df.drop('CuentaTemp', axis=1, inplace=True)
    #print(df)
    cuentas_129 = [x for x in df.cuenta if x[:3] in [str(y) for y in dictionaries.conceptosAbreviado['PASIVO']['resultado del ejercicio']]]
    #print('Cuenta 129: ', cuentas_129)
    if not cuentas_129:
        cuenta = '129' + '0' * (max - 3)
        df.loc[len(df)] = [cuenta, 'resultado del ejercicio', 0]
        cuentas_129 = [cuenta]

    insert_variacion = True
    cuentas = list(set([str(x)[0:2] for x in df.cuenta]))
    cuentas += list(set([str(x)[0:4] for x in df.cuenta]))
    #print(cuentas)
    #print([x for x in cuentas if x in ['61', '71', '6930', '7930']])
    if [x for x in cuentas if
        x in [str(y) for y in dictionaries.conceptosAbreviado['RESULTADOS']['variacion de existencias de productos terminados y en curso de fabricacion']]]:
        insert_variacion = False

    if insert_variacion and conceptos['variacion']:
        #print(sign, variacion)
        cuenta = '300' + '0' * (max - 3)
        #i = 0
        #while not df[df['cuenta'] == cuenta].empty:
        #    i += 1
        #    cuenta = str(int(cuenta + i))
        df.loc[len(df)] = [cuenta, 'correccion existencias activo', conceptos['variacion'] * sign * (-1.0)]

        cuenta = '610' + '0' * (max - 3)
        df.loc[len(df)] = [cuenta, 'correccion existencias resultados', conceptos['variacion'] * sign]

    insert_amortizacion = True
    cuentas = list(set([str(x)[0:2] for x in df.cuenta]))
    if '68' in cuentas:
        insert_amortizacion = False

    if insert_amortizacion and conceptos['amortizacion']:
        #print(sign, variacion)
        cuenta = '680' + '0' * (max - 3)
        df.loc[len(df)] = [cuenta, 'correccion amortizaciones resultados', conceptos['amortizacion'] * sign]

    extraccion.ruta_unmerged = None
    extraccion.save()

    os.remove(excel)
    book.close()
    #print(df)
    #print(cuentas_129)

    return df, cuentas_129

"""
if __name__ == '__main__':
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    desired_width = 2500
    pd.set_option('display.width', desired_width)

    df = read('C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\SumasSaldos\\SUMAS Y SALDO 31-10-2021.xlsx',
         'C:\\Users\\Pablo\\OneDrive\\Tecnocim\\FicherosAlia\\SumasSaldos\\SUMAS Y SALDO 31-10-2021_unmerged.xlsx')
    print(df)
"""